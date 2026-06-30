"""Tests for export_to_excel days= filter."""
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from db.base import Base
from db.models import NewsCard, RawPost, Source


@pytest.fixture
def mem_db(tmp_path):
    """In-memory SQLite with patched get_session for the exporter."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    _Session = sessionmaker(engine, autocommit=False, autoflush=False)

    @contextmanager
    def _get_session():
        s = _Session()
        try:
            yield s
            s.commit()
        except Exception:
            s.rollback()
            raise
        finally:
            s.close()

    with patch("exporter.excel.get_session", _get_session):
        yield engine, _get_session, tmp_path


def _insert_card(get_session_fn, days_old: float, source_suffix: str = "") -> int:
    """Create Source → RawPost → NewsCard with both dates set to `days_old` days ago."""
    pub = datetime.utcnow() - timedelta(days=days_old)
    with get_session_fn() as s:
        src = Source(
            username=f"@test_ch{source_suffix}",
            title=f"Test Channel{source_suffix}",
            topics='["тест"]',
            is_active=True,
        )
        s.add(src)
        s.flush()
        post = RawPost(
            source_id=src.id,
            message_id=abs(hash(source_suffix)) % 100000,
            channel_username=f"@test_ch{source_suffix}",
            post_url=f"https://t.me/test{source_suffix}/1",
            raw_text="Тестовый текст новости",
            text_hash=f"hash{source_suffix}",
            published_at=pub,
            has_media=False,
            is_forwarded=False,
        )
        s.add(post)
        s.flush()
        card = NewsCard(
            raw_post_id=post.id,
            title=f"Card {source_suffix}",
            source_title=f"Test Channel{source_suffix}",
            post_url=f"https://t.me/test{source_suffix}/1",
            published_at=pub,
            created_at=pub,
            clean_text="clean text",
            relevance_score=50,
            relevance_label="medium",
        )
        s.add(card)
        s.flush()
        return card.id


class TestExcelDaysFilter:
    def test_no_filter_exports_all_cards(self, mem_db):
        engine, gs, tmp_path = mem_db
        _insert_card(gs, days_old=15, source_suffix="_15d")
        _insert_card(gs, days_old=1, source_suffix="_1d")

        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=str(tmp_path / "all.xlsx"), days=None)

        wb = openpyxl.load_workbook(path)
        ws = wb["news_cards"]
        data_rows = ws.max_row - 1  # row 1 is header
        assert data_rows == 2

    def test_days_filter_excludes_old_cards(self, mem_db):
        engine, gs, tmp_path = mem_db
        _insert_card(gs, days_old=15, source_suffix="_old")
        _insert_card(gs, days_old=1, source_suffix="_new")

        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=str(tmp_path / "filtered.xlsx"), days=7)

        wb = openpyxl.load_workbook(path)
        ws = wb["news_cards"]
        data_rows = ws.max_row - 1
        assert data_rows == 1

    def test_days_filter_card_on_boundary_is_included(self, mem_db):
        engine, gs, tmp_path = mem_db
        _insert_card(gs, days_old=6, source_suffix="_boundary")

        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=str(tmp_path / "boundary.xlsx"), days=7)

        wb = openpyxl.load_workbook(path)
        ws = wb["news_cards"]
        data_rows = ws.max_row - 1
        assert data_rows == 1

    def test_all_old_cards_excluded(self, mem_db):
        engine, gs, tmp_path = mem_db
        _insert_card(gs, days_old=30, source_suffix="_30d")
        _insert_card(gs, days_old=20, source_suffix="_20d")

        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=str(tmp_path / "empty.xlsx"), days=7)

        wb = openpyxl.load_workbook(path)
        ws = wb["news_cards"]
        data_rows = ws.max_row - 1
        assert data_rows == 0

    def test_output_file_is_created(self, mem_db):
        engine, gs, tmp_path = mem_db
        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=str(tmp_path / "out.xlsx"), days=7)
        assert Path(path).exists()

    def test_returned_path_matches_output_path(self, mem_db):
        engine, gs, tmp_path = mem_db
        out = str(tmp_path / "specific.xlsx")
        from exporter.excel import export_to_excel
        path = export_to_excel(output_path=out, days=7)
        assert path == out

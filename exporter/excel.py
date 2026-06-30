import json
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import or_, select

from db.base import get_session
from db.models import NewsCard, RawPost, Trend, TrendCase

# --- Column definitions -------------------------------------------------------

_CARDS_COLS = [
    "Дата", "Источник", "Заголовок", "Ссылка на пост",
    "Внешняя ссылка", "Резюме для сайта",
    "Темы", "Теги", "Компании", "Релевантность", "Уровень",
    "LLM релевантность", "Реклама", "Статус проверки", "Обогащён", "Текст",
]

_CARDS_WIDTHS: dict[str, int] = {
    "A": 18, "B": 20, "C": 45, "D": 12,
    "E": 14, "F": 60, "G": 30, "H": 35,
    "I": 25, "J": 14, "K": 12, "L": 16,
    "M": 10, "N": 16, "O": 12, "P": 60,
}

_POSTS_COLS = [
    "ID", "Канал", "Message ID", "Дата публикации", "Ссылка",
    "Есть медиа", "Репост", "Просмотры", "Обработан", "Текст",
]

_POSTS_WIDTHS: dict[str, int] = {
    "A": 6,  "B": 20, "C": 12, "D": 18, "E": 12,
    "F": 12, "G": 10, "H": 12, "I": 12, "J": 60,
}

_CASES_COLS = [
    "Дата", "Источник", "Тренд", "Кейс", "Компания",
    "Описание", "Как работает", "Ценность", "Рынок", "Ссылка", "Заголовок поста",
    "Тренд ID", "Период",
]

_CASES_WIDTHS: dict[str, int] = {
    "A": 18, "B": 20, "C": 30, "D": 40, "E": 25,
    "F": 60, "G": 50, "H": 40, "I": 15, "J": 14, "K": 45,
    "L": 8, "M": 10,
}

_TRENDS_COLS = ["ID", "Тренд", "Описание", "Кейсов", "Первый кейс", "Обновлён"]

_TRENDS_WIDTHS: dict[str, int] = {
    "A": 6, "B": 35, "C": 60, "D": 10, "E": 15, "F": 15,
}

_BY_CHANNEL_COLS = ["Канал", "Постов", "Кейсов", "Средний score", "Рекламных постов"]

_BY_CHANNEL_WIDTHS: dict[str, int] = {
    "A": 35, "B": 10, "C": 10, "D": 14, "E": 18,
}

_COLOR_AD = "FFEB9C"
_COLOR_HIGH = "C6EFCE"
_COLOR_LOW = "F2F2F2"
_COLOR_HEADER = "D9D9D9"

# --- Helper functions ---------------------------------------------------------


def _json_to_str(value: str | None) -> str:
    """JSON-строка ['a', 'b'] → 'a, b'. При ошибке возвращает value as-is."""
    if not value:
        return ""
    try:
        items = json.loads(value)
        if isinstance(items, list):
            return ", ".join(str(x) for x in items)
        return str(items)
    except (json.JSONDecodeError, TypeError):
        return value or ""


def _bool_to_ru(value: bool) -> str:
    """True → 'Да', False → ''"""
    return "Да" if value else ""


def _llm_bool_to_ru(value: bool | None) -> str:
    """True → 'Да', False → 'Нет', None → '—'"""
    if value is True:
        return "Да"
    if value is False:
        return "Нет"
    return "—"


def _fmt_dt(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y %H:%M")


def _apply_header_style(ws: Worksheet, num_cols: int) -> None:
    """Жирный шрифт + серый фон для строки 1."""
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor=_COLOR_HEADER)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill


def _apply_row_color(ws: Worksheet, row_num: int, num_cols: int, color: str) -> None:
    """Закрасить строку row_num цветом color (hex без #)."""
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, num_cols + 1):
        ws.cell(row=row_num, column=col).fill = fill


def _set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    """Принимает {letter: width} и устанавливает ширины."""
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _apply_hyperlinks(
    ws: Worksheet,
    col_idx: int,
    data_rows: list[dict],
    url_key: str = "_url",
    link_text: str = "Открыть",
) -> None:
    for i, row in enumerate(data_rows, start=2):
        url = row.get(url_key) or ""
        cell = ws.cell(row=i, column=col_idx)
        cell.value = f'=HYPERLINK("{url}", "{link_text}")' if url else ""


def _format_sheet(
    ws: Worksheet,
    data_rows: list[dict],
    widths: dict[str, int],
    url_col_idx: int | None = None,
    apply_card_colors: bool = False,
    source_url_col_idx: int | None = None,
    wrap_col_idxs: list[int] | None = None,
    wrap_height: int = 80,
) -> None:
    num_cols = ws.max_column or len(widths)
    _apply_header_style(ws, num_cols)
    _set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}{max(ws.max_row, 1)}"
    if url_col_idx is not None:
        _apply_hyperlinks(ws, url_col_idx, data_rows)
    if source_url_col_idx is not None:
        _apply_hyperlinks(ws, source_url_col_idx, data_rows, "_source_url", "Источник")

    for i, row in enumerate(data_rows, start=2):
        if apply_card_colors:
            is_ad = row.get("_is_ad", False)
            label = row.get("_relevance_label", "")
            if is_ad:
                row_color = _COLOR_AD
            elif label == "high":
                row_color = _COLOR_HIGH
            elif label in ("low", "irrelevant"):
                row_color = _COLOR_LOW
            else:
                row_color = None
            if row_color:
                _apply_row_color(ws, i, num_cols, row_color)
        if wrap_col_idxs:
            for col_idx in wrap_col_idxs:
                ws.cell(row=i, column=col_idx).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[i].height = wrap_height


# --- Row builders -------------------------------------------------------------


def _card_to_row(card: NewsCard) -> dict:
    return {
        "Дата": _fmt_dt(card.published_at),
        "Источник": card.source_title or "",
        "Заголовок": card.title or "",
        "Ссылка на пост": "Открыть",
        "Внешняя ссылка": "Источник" if card.source_url else "",
        "Резюме для сайта": card.summary or "",
        "Темы": _json_to_str(card.topics),
        "Теги": _json_to_str(card.tags),
        "Компании": _json_to_str(card.companies),
        "Релевантность": card.relevance_score,
        "Уровень": card.relevance_label,
        "LLM релевантность": _llm_bool_to_ru(card.llm_relevant),
        "Реклама": _bool_to_ru(card.is_ad),
        "Статус проверки": card.review_status,
        "Обогащён": _bool_to_ru(card.llm_enriched),
        "Текст": card.clean_text or "",
        # private fields — not written to DataFrame, used only for formatting
        "_url": card.post_url or "",
        "_source_url": card.source_url or "",
        "_is_ad": card.is_ad,
        "_relevance_label": card.relevance_label,
    }


def _post_to_row(post: RawPost, has_card: bool) -> dict:
    return {
        "ID": post.id,
        "Канал": post.channel_username,
        "Message ID": post.message_id,
        "Дата публикации": _fmt_dt(post.published_at),
        "Ссылка": "Открыть",
        "Есть медиа": _bool_to_ru(post.has_media),
        "Репост": _bool_to_ru(post.is_forwarded),
        "Просмотры": post.views if post.views is not None else "",
        "Обработан": "Да" if has_card else "Нет",
        "Текст": post.raw_text or "",
        "_url": post.post_url,
    }


def _case_to_row(case: TrendCase) -> dict:
    nc = case.news_card
    return {
        "Дата": _fmt_dt(nc.published_at) if nc else "",
        "Источник": nc.source_title or "" if nc else "",
        "Тренд": case.trend_name or "",
        "Кейс": case.case_title or "",
        "Компания": case.company or "",
        "Описание": case.description or "",
        "Как работает": case.how_it_works or "",
        "Ценность": case.value or "",
        "Рынок": case.market or "",
        "Ссылка": "Открыть" if case.source_url else "",
        "Заголовок поста": nc.title or "" if nc else "",
        "Тренд ID": case.trend_id or "",
        "Период": case.period_label or "",
        "_url": case.source_url or "",
    }


def _trend_to_row(trend: Trend, cases_count: int) -> dict:
    return {
        "ID": trend.id,
        "Тренд": trend.name,
        "Описание": trend.description or "",
        "Кейсов": cases_count,
        "Первый кейс": _fmt_dt(trend.first_seen_at),
        "Обновлён": _fmt_dt(trend.updated_at),
    }


def _rows_to_df(rows: list[dict], cols: list[str]) -> pd.DataFrame:
    public = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
    return pd.DataFrame(public, columns=cols)


def _build_by_channel_rows(
    cards_rows: list[dict], cases_rows: list[dict]
) -> list[dict]:
    from collections import defaultdict

    stats: dict = defaultdict(lambda: {"posts": 0, "cases": 0, "scores": [], "ads": 0})
    for row in cards_rows:
        ch = row.get("Источник") or "—"
        stats[ch]["posts"] += 1
        score = row.get("Релевантность")
        if score:
            stats[ch]["scores"].append(score)
        if row.get("Реклама"):
            stats[ch]["ads"] += 1
    for row in cases_rows:
        ch = row.get("Источник") or "—"
        stats[ch]["cases"] += 1

    return [
        {
            "Канал": ch,
            "Постов": s["posts"],
            "Кейсов": s["cases"],
            "Средний score": (
                round(sum(s["scores"]) / len(s["scores"]), 1) if s["scores"] else 0
            ),
            "Рекламных постов": s["ads"],
        }
        for ch, s in sorted(stats.items(), key=lambda x: x[1]["cases"], reverse=True)
    ]


# --- Main export function -----------------------------------------------------


def export_to_excel(output_path: str | None = None, days: int | None = None) -> str:
    """
    Экспортирует данные из БД в Excel.
    days — если задан, фильтрует news_cards, raw_posts и trend_cases за последние N дней.
    output_path — если None, генерирует имя автоматически.
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"data/exports/news_export_{ts}.xlsx"

    since: datetime | None = None
    if days is not None:
        since = datetime.utcnow() - timedelta(days=days)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Starting export → {output_path}" + (f" (last {days}d)" if days else ""))

    with get_session() as session:
        cards_stmt = select(NewsCard).order_by(NewsCard.id)
        if since:
            cards_stmt = cards_stmt.where(
                or_(NewsCard.published_at >= since, NewsCard.created_at >= since)
            )
        cards = session.execute(cards_stmt).scalars().all()

        posts_stmt = select(RawPost).order_by(RawPost.id)
        if since:
            posts_stmt = posts_stmt.where(RawPost.published_at >= since)
        posts = session.execute(posts_stmt).scalars().all()

        processed_ids: set[int] = set(
            session.execute(select(NewsCard.raw_post_id)).scalars().all()
        )

        review_stmt = (
            select(NewsCard)
            .where(
                NewsCard.review_status.in_(["auto", "needs_review"]),
                NewsCard.publish_status == "draft",
            )
        )
        if since:
            review_stmt = review_stmt.where(
                or_(NewsCard.published_at >= since, NewsCard.created_at >= since)
            )
        review_stmt = review_stmt.order_by(
            NewsCard.relevance_score.desc(), NewsCard.published_at.desc()
        )
        review_cards = session.execute(review_stmt).scalars().all()

        cases_stmt = select(TrendCase).order_by(TrendCase.id)
        if since:
            cases_stmt = cases_stmt.where(TrendCase.created_at >= since)
        cases = session.execute(cases_stmt).scalars().all()

        all_trends = session.execute(
            select(Trend).order_by(Trend.first_seen_at.desc())
        ).scalars().all()

        trend_case_counts = Counter(c.trend_id for c in cases if c.trend_id is not None)

        # Build rows while session is still open — objects are still attached.
        cards_rows = [_card_to_row(c) for c in cards]
        posts_rows = [_post_to_row(p, p.id in processed_ids) for p in posts]
        review_rows = [_card_to_row(c) for c in review_cards]
        cases_rows = [_case_to_row(c) for c in cases]
        trends_rows = [_trend_to_row(t, trend_case_counts.get(t.id, 0)) for t in all_trends]

    by_channel_rows = _build_by_channel_rows(cards_rows, cases_rows)

    logger.info(
        f"Rows: news_cards={len(cards_rows)},"
        f" raw_posts={len(posts_rows)}, review={len(review_rows)},"
        f" trend_cases={len(cases_rows)}, trends={len(trends_rows)},"
        f" by_channel={len(by_channel_rows)}"
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _rows_to_df(cards_rows, _CARDS_COLS).to_excel(
            writer, sheet_name="news_cards", index=False
        )
        _rows_to_df(posts_rows, _POSTS_COLS).to_excel(
            writer, sheet_name="raw_posts", index=False
        )
        _rows_to_df(review_rows, _CARDS_COLS).to_excel(
            writer, sheet_name="review", index=False
        )
        _rows_to_df(cases_rows, _CASES_COLS).to_excel(
            writer, sheet_name="trend_cases", index=False
        )
        _rows_to_df(trends_rows, _TRENDS_COLS).to_excel(
            writer, sheet_name="trends", index=False
        )
        _rows_to_df(by_channel_rows, _BY_CHANNEL_COLS).to_excel(
            writer, sheet_name="by_channel", index=False
        )

        _format_sheet(
            writer.sheets["news_cards"],
            cards_rows,
            _CARDS_WIDTHS,
            url_col_idx=4,
            source_url_col_idx=5,
            wrap_col_idxs=[6, 16],
            wrap_height=80,
            apply_card_colors=True,
        )
        _format_sheet(
            writer.sheets["raw_posts"],
            posts_rows,
            _POSTS_WIDTHS,
            url_col_idx=5,
            wrap_col_idxs=[10],
            wrap_height=60,
        )
        _format_sheet(
            writer.sheets["review"],
            review_rows,
            _CARDS_WIDTHS,
            url_col_idx=4,
            source_url_col_idx=5,
            wrap_col_idxs=[6, 16],
            wrap_height=80,
            apply_card_colors=True,
        )
        _format_sheet(
            writer.sheets["trend_cases"],
            cases_rows,
            _CASES_WIDTHS,
            url_col_idx=10,
            wrap_col_idxs=[6, 7, 8],
            wrap_height=80,
        )
        _format_sheet(
            writer.sheets["trends"],
            trends_rows,
            _TRENDS_WIDTHS,
            wrap_col_idxs=[3],
            wrap_height=80,
        )
        _format_sheet(
            writer.sheets["by_channel"],
            by_channel_rows,
            _BY_CHANNEL_WIDTHS,
        )

    logger.info(f"Export complete: {output_path}")
    return output_path
